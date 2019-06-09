import openpyxl


def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws["B2"] = "Ticker"
    ws["C2"] = "Pct Change"

    row, col = 3, 2

    ticker_changes = [
        ("AAPL", -0.03),
        ("TSLA", 0.021),
        ("FB", 0.0),
        ("NFLX", -0.05),
        ("GOOGL", 0.032)
    ]
    for tc in ticker_changes:
        ws.cell(row, col).value = tc[0]
        ws.cell(row, col + 1).value = tc[1]
        row += 1

    # Conditional formatting
    from openpyxl.styles import Color, PatternFill, Font, Border
    from openpyxl.formatting.rule import CellIsRule

    def pct_red_green_fmt(ref):
        redFill = PatternFill(start_color="EE1111",
                              end_color="EE1111", fill_type="solid")
        greenFill = PatternFill(start_color="007700",
                                end_color="007700", fill_type="solid")
        ws.conditional_formatting.add(ref, CellIsRule(
            operator="lessThan", formula=[0], stopIfTrue=False, fill=redFill))
        ws.conditional_formatting.add(ref, CellIsRule(
            operator="greaterThan", formula=[0], stopIfTrue=False, fill=greenFill))

    # Create table
    from openpyxl.worksheet.table import Table

    def to_table(ref, headers=True):
        tab = Table(displayName="Table1", ref=ref)
        ws.add_table(tab)

    pct_red_green_fmt('C3:C7')
    to_table('B2:C7')
    wb.save("workbooks/mywb.xlsx")


if __name__ == "__main__":
    main()
